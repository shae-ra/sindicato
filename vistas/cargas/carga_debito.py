#=============
#IMPORTACIONES
#=============

# Importamos el módulo sys que provee el acceso a funciones y objetos mantenidos por el intérprete.
import sys
# Importamos las herramientas de PyQT que vamos a utilizar
from PyQt5 import QtWidgets, uic, QtGui
# Importamos los elementos que se encuentran dentro del diseñador
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QPushButton, QTabWidget, QMessageBox
from PyQt5.QtCore import Qt
# Importamos el modulo uic necesario para levantar un archivo .ui
from PyQt5 import uic
import openpyxl
import win32print
from configparser import ConfigParser

from datetime import date

from modelos.modelo_proveedor import ModeloProveedor
from modelos.modelo_debito import ModeloDebito

#====================
#DEFINICION DE CLASES
#====================

#Creacion de la clase CargaDebito
class CargaDebito(QtWidgets.QWidget):
	#Inicializacion del Objeto QWidget
	def __init__(self, parent = None):
		QWidget.__init__(self, parent = None)

		self.parent = parent
		# Importamos la vista "carga_debito" y la alojamos dentro de la variable "carga"
		self.v_carga = uic.loadUi("gui/cargas/carga_debito.ui", self)

		self.model_prov = ModeloProveedor(propiedades = ['id', 'nombre'])
		self.model = ModeloDebito()

		self.v_carga.prov_id.setModel(self.model_prov)
		self.v_carga.prov_id.currentIndexChanged.connect(self.__disableComision)
		self.v_carga.prov_id.currentIndexChanged.connect(self.setNumeroDeOrden)

		self.v_carga.btn_confirmar.clicked.connect(self.guardarDebito)

		self.v_carga.deb_total_cuotas.textChanged.connect(self.__calcularTotalACobrar)
		self.v_carga.deb_importe_cuota.textChanged.connect(self.__calcularTotalACobrar)


	def guardarDebito(self):
		debito = self.getDebito()
		if debito['errores']:
			self.mensajeError(debito['errores'])
			#observer.msg("No se puede cargar el debito")
		else:
			debito.pop('errores')
			of = debito['importe_actual'] % 900
			fecha = debito['fecha_descuento']
			for subDebito in range(int(debito['importe_actual']/900)):
				overflow = self.overflowDebito(debito)
				self.model.guardarDebito(overflow)
				debito['fecha_descuento'] = fecha
			debito['importe_actual'] = of
			if debito['importe_actual'] > 0:
				self.model.guardarDebito(debito)
			self.getXls()
			self.operacionCompletada()
			reset = self.resetDebito()
			self.close()
		return

	def mensajeError(self, errores):
		msg = QMessageBox()
		msg.setIcon(QMessageBox.Information)
		erroresMsg = ""
		for error in errores:
			erroresMsg += error
		msg.setText("No se puede realizar la operacion por los siguientes motivos: \n" + erroresMsg )
		msg.setWindowTitle("...")
		msg.exec_()

	def operacionCompletada(self):
		msg = QMessageBox()
		msg.setIcon(QMessageBox.Information)
		msg.setText("Operación Exitosa     ")
		msg.setWindowTitle("...")
		retval = msg.exec_()

	def resetDebito(self):
		self.v_carga.deb_fecha_mes.setText('')
		self.v_carga.deb_fecha_anio.setText('')
		self.v_carga.deb_total_cuotas.setText('')
		self.v_carga.deb_importe_cuota.setText('')
		self.v_carga.deb_importe_total.setText('')
		self.v_carga.deb_orden.setText('')

	def getDebito(self):
		errores = []
		debito = {
		"legajo_afiliado" : self.parent.vd_afiliado.af_legajo.text(),
		"fecha_carga_inicial" : date.today(),
		"n_orden" : self.v_carga.deb_orden.text()
		}

		fecha_mes = 0
		fecha_anio = 0
		try:
			fecha_anio = int(self.v_carga.deb_fecha_anio.text())
		except:
			errores.append("- El año ingresado no es válido\n")
		try:
			fecha_mes = int(self.v_carga.deb_fecha_mes.text())
		except:
			errores.append("- El mes ingresado no es válido\n")
		try:
			debito['proveedor_id'] = int(self.v_carga.prov_id.currentText().split("-")[0])
		except:
			errores.append("- No hay seleccionado un proveedor\n")
		try:
			debito['total_cuotas'] = int(self.v_carga.deb_total_cuotas.text())
			if debito['total_cuotas'] == 0:
				errores.append("- No hay cuotas ingresadas\n")
		except:
			errores.append("- No hay cuotas ingresadas\n")
		try:
			debito['importe_total'] = int(self.v_carga.deb_importe_total.text())
			if debito['importe_total'] == "" or debito['importe_total'] == "0":
				errores.append("- No se ha podido calcular el total a cobrar\n")
		except:
			errores.append("- No se ha podido calcular el total a cobrar\n")

		if fecha_mes and fecha_anio:
			debito['fecha_descuento'] = date(fecha_anio, fecha_mes, 1)
		try:
			debito['importe_actual'] = int(self.v_carga.deb_importe_cuota.text())
		except:
			errores.append("- No se ha ingresado un importe\n")

		# print(debito)
		debito['errores'] = errores
		return debito

	def getXls(self):

		wb = openpyxl.load_workbook('bono/bonos.xlsx')
		ws = wb.worksheets[0]
		img = openpyxl.drawing.image.Image('bono/sindicato.png')
		img2 = openpyxl.drawing.image.Image('bono/sindicato.png')
		img3 = openpyxl.drawing.image.Image('bono/sindicato.png')
		ws.add_image(img, 'A1')
		ws.add_image(img2, 'A15')
		ws.add_image(img3, 'A29')

		ws['C4'] = date.today()
		ws['D5'] = "{} {}".format(self.parent.vd_afiliado.af_apellido.text(), self.parent.vd_afiliado.af_nombre.text())
		ws['C6'] = self.parent.vd_afiliado.af_legajo.text()
		ws['E6'] = self.parent.vd_afiliado.af_lugar_trabajo.text()
		ws['C7'] = self.v_carga.deb_orden.text()
		ws['E7'] = self.v_carga.prov_id.currentText().split("-")[1]
		ws['H9'] = self.v_carga.deb_importe_total.text()
		ws['D9'] = self.v_carga.deb_importe_palabras.text()

		# ABRIR UN CUADRO DE DIALOGO INDICANDO DONDE GUARDAR
		self.handleSave(wb)

		wb.close()

	def handleSave(self, workbook):
		path = QtWidgets.QFileDialog.getSaveFileName(
			None, 'Save File', self.v_carga.deb_orden.text(), 'Excel(*.xlsx)')
		if not path[0]: return
		workbook.save(path[0])

	def imprimirBono(self):
		#TERMINAR ESTO
		config = ConfigParser()

		if config.read('impresora.ini'):
			if 'IMPRESORA' in config.section():
				impresora = config['IMPRESORA']['nombre_impresora']
				printer = win32print.SetDefaultPrinter(impresora)

		else:
			print("No se ha configurado la impresora")

	def overflowDebito(self, debito):
		debito['importe_actual'] = 900
		return debito

	def showEvent(self, event):
		self.model_prov.verListaProveedores()
		self.v_carga.prov_id.setCurrentIndex(0)

	def setNumeroDeOrden(self):
		numeroDeOrden = self.getNumeroDeOrden()

		self.v_carga.deb_orden.setText(numeroDeOrden)

	def getNumeroDeOrden(self):
		proveedor = int(self.v_carga.prov_id.currentText().split("-")[0])
		fecha = date.today()

		numero = self.model.consultarUltimoNumeroDeOrden(proveedor, fecha)
		numero = str(int(numero) + 1)

		fecha = self.__formatearFecha(fecha)
		proveedor = self.__formatearIdProveedor(proveedor)
		numero = self.__formatearNumeroOrden(numero)


		numeroDeOrden = fecha + proveedor + numero
		return numeroDeOrden

	def __formatearFecha(self, fecha):
		fecha = fecha.strftime('%d%m%Y')
		return fecha

	def __formatearIdProveedor(self, idProveedor):
		idProveedor = str(idProveedor)
		while len(idProveedor) < 2:
			idProveedor = '0' + idProveedor
		return idProveedor

	def __formatearNumeroOrden(self, numero):
		numero = str(numero)
		while len(numero) < 2:
			numero = '0' + numero
		return numero

	def __calcularTotalACobrar(self):
		cantidad_cuotas = 0
		importe = 0
		try:
			cantidad_cuotas = int(self.v_carga.deb_total_cuotas.text())
			importe = int(self.v_carga.deb_importe_cuota.text())
		except ValueError:
			total_a_cobrar = 0
		total_a_cobrar = cantidad_cuotas * importe

		self.v_carga.deb_importe_total.setText(str(total_a_cobrar))

	def __disableComision(self):
		if int(self.v_carga.prov_id.currentText().split("-")[0]) == 7:
			self.v_carga.deb_orden.setEnabled(True)
		else:
			self.v_carga.deb_orden.setEnabled(False)

	def keyPressEvent(self, event):
		if event.key() == Qt.Key_Escape:
			self.close()
