from PyQt5 import QtCore
from libs.db import querier
from datetime import date
from dateutil.relativedelta import relativedelta
from PyQt5.QtWidgets import QFileDialog
import openpyxl
import cerberus

class ModeloProcesador(QtCore.QAbstractTableModel):
    __querier = querier.Querier()
    __v = cerberus.Validator()

        # La idea ahora es separar el esquema de validación de los datos y campos que se van a usar, nosotros no vamos a usar
        # todos los campos en la tabla, habíamos definido los que se encuentran en el archivo 'cosas para hacer.md'
        # | Mes | Proveedor | Importe | Cuota N° | Total Cuotas

    def __init__(self, propiedades = None, parent = None):
        super(ModeloProcesador, self).__init__()
        if parent:
            self.__parent = parent
        self.__propiedades = ["Estado", "Fecha", "Legajo", "Banco", "CBU", "Importe", "CUIT", "Orden de movimiento", "Codigo de rechazo", "Motivo de rechazo", "Empresa"
        ]

        if propiedades:
            self.__propiedades = self.validarPropiedades(propiedades)

        self.__debitosBet = [] # Los valores de prueba los saco del archivo fuente

        self.__debitosDatabase = [] # Los valores de acá salen de la base de datos
        self.__debitosNoProcesados = []
        self.__debitosRechazados = []
        self.__debitosProcesados = []
        self.__codigosDeRechazo = {}

        self.verListaCodigosDeRechazo()

        # self.setConfig(banco = "P", cuit = "30561600194", empresa = "SIND T MUN MERLO")

    def verListaCodigosDeRechazo(self):

        self.__codigosDeRechazo = self.__querier.traerElementos(
            campos = ["codigo", "descripcion"],
            tabla = "codigos_rechazo"
        )
        codigosDeRechazo = dict( (codigo, descripcion) for codigo, descripcion in self.__codigosDeRechazo)
        self.__codigosDeRechazo = codigosDeRechazo
        self.__codigosDeRechazo[None] = ""
        self.__codigosDeRechazo['   '] = ""
        # CodigosDeRechazo ahora contiene una lista de tuplas, cada tupla es un registro de la base de datos.

    def verListaDebitosAProcesar(self, lista):

        self.__debitosBet = lista
        self.verListaDebitosDatabase()
        self.__debitosProcesados = []
        self.__debitosProcesadosDB = []
        self.__debitosRechazados = []
        # self.procesables = []

        for index, debito in enumerate(self.__debitosBet):
            codigo = debito[8]

            if codigo != "   ":
                self.__debitosBet[index][9] = self.__codigosDeRechazo[codigo]

            self.obtenerDebitoProcesable(debito)

        # if self.procesables:
        print("DEBUG - Debitos a procesar: ", self.__debitosBet)
        print("DEBUG - Procesados: ", self.__debitosProcesados)
        print("DEBUG - Rechazados, para clonar: ", self.__debitosRechazados)
        print("DEBUG - No Procesados, para clonar: ", self.__debitosNoProcesados)
            # self.__debitosBet = self.procesables
        self.layoutChanged.emit()
        return True
        # return False

    def verListaDebitosDatabase(self):

        self.__debitosDatabase = self.__querier.traerElementos(
            campos = ["id","id_temporal", "legajo_afiliado", "cbu", "fecha_descuento", "importe_actual",
                "fecha_carga_inicial", "proveedor_id", "cuota_actual", "total_cuotas", "importe_total", "n_orden", "estado", "motivo"],
            tabla = "debitos",
            uniones = [("afiliados", "afiliados.legajo = debitos.legajo_afiliado")],
            condiciones = [("id_temporal", "IS NOT", "NULL")],
            orden = ("id_temporal", "ASC")
        )

        print(self.__debitosDatabase)

    def apllicarCambios(self):
        self.guardarXls()
        for debito in self.__debitosProcesados:
            print("Estado a actualizar: " + debito[0])
            print("Motivo a actualizar: " + debito[8])
            self.__querier.actualizarElemento(
                tabla = "debitos",
                elemento = { "id_temporal" : None,
                    "estado" : debito[0],
                    "motivo" : debito[8] },
                condiciones = [("id_temporal", "=", int(debito[7])), ("legajo_afiliado", "=", debito[2])]
            )

        for debito in self.__debitosRechazados:
            self.__querier.insertarElemento(
                tabla = "debitos",
                elemento = { "id_temporal" : None,
                    "legajo_afiliado" : debito[2],
                    "fecha_descuento" : debito[4] + relativedelta(months = 1), # Nueva fecha para proximo mes
                    "importe_actual" : debito[5],
                    "fecha_carga_inicial" : debito[6],
                    "proveedor_id" : debito[7],
                    "cuota_actual" : debito[8],
                    "total_cuotas" : debito[9],
                    "importe_total" : debito[10],
                    "n_orden" : debito[11],
                    "estado" : None,
                    "motivo" : None }
            )
            self.__querier.actualizarElemento(
                tabla = "debitos",
                elemento = { "id_temporal" : None,
                    "estado" : debito[12],
                    "motivo" : debito[13]
                    },
                condiciones = [("id", "=", debito[0])]
            )

        self.verListaDebitosDatabase()
        self.limpiarTabla()

    def actualizarDebito(self, debito, condiciones):
        self.__querier.actualizarElemento(
            tabla = "debitos",
            elemento = debito,
            condiciones = condiciones
        )

    def obtenerDebitoProcesable(self, debito):
        # Voy a comparar los elementos de self.__debitosBet con los de
        # self.__debitosDatabase. El segundo lo puedo traer ordenado para hacer de las búsquedas un proceso mas rápido,
        # el primero no hace falta ordenarlo ya que vamos a iterar sobre el registro por registro.

        # noProcesados = []
        # ignorados = []

        match = []

        f_idTemporal, f_legajoAfiliado, f_cbu, f_fecha, f_importe, f_estado, f_codigo_error = 7, 2, 4, 1, 5, 0, 8
        db_idTemporal, db_legajoAfiliado, db_cbu, db_fecha, db_importe, db_estado, db_motivo = 1, 2, 3, 4, 5, 12, 13

        for index, possMatch in enumerate(self.__debitosDatabase):
            if possMatch[db_idTemporal] == int(debito[f_idTemporal]) and possMatch[db_legajoAfiliado] == debito[f_legajoAfiliado]:
                match = self.__debitosDatabase[index]

                print("\nDEBUG - El item file contiene: ")
                print("Legajo de Afiliado: ", debito[f_legajoAfiliado])
                print("Id Temporal: ", debito[f_idTemporal])
                print("Estado: ", debito[f_estado])
                print("Codigo de error: ", debito[f_codigo_error])

                print("\nDEBUG - El item en db contiene: ")
                print("Legajo de Afiliado: ", match[db_legajoAfiliado])
                print("Id Temporal", match[db_idTemporal])

                print("")

                if debito[f_cbu] != match[db_cbu]: # CBU
                    print("Los CBU no coinciden en ", match[db_legajoAfiliado])
                    print(debito[f_cbu])
                    print(match[db_cbu])
                if debito[f_fecha] != match[db_fecha]:  # Fecha
                    print("Las fechas no coinciden en ", match[db_legajoAfiliado])
                    print(debito[f_fecha])
                    print(match[db_fecha].strftime('%d/%m/%Y'))
                if debito[f_importe] != match[db_importe]:  # Importe
                    print("Los importes no coinciden en ", match[db_legajoAfiliado])
                    print(debito[f_importe])
                    print(match[db_importe])

                if debito[f_estado] != "Procesado" or debito[f_codigo_error] != "   ":
                    match = list(match)
                    match[db_estado] = debito[f_estado]
                    match[db_motivo] = debito[f_codigo_error]
                    self.__debitosRechazados.append(match)
                else:
                    self.__debitosProcesadosDB.append(list(match))
                    self.__debitosProcesados.append(list(debito))
                return True
        return False

    def guardarXls(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        registrosProcesados = len(self.__debitosProcesados)
        registrosRechazados = len(self.__debitosRechazados)
        # registrosModificados = registrosProcesados + registrosRechazados

        debitosXls = self.__querier.traerElementos(
            tabla = "debitos",
            campos = ["CONCAT(afiliados.nombre, ' ', afiliados. apellido)", "importe_actual", "cuota_actual", "total_cuotas", "motivo", "proveedores.nombre", "id_temporal", "legajo_afiliado"],
            uniones = [('afiliados', 'afiliados.legajo = debitos.legajo_afiliado'), ('proveedores', 'proveedores.id = debitos.proveedor_id')],
            condiciones = [('id_temporal','IS NOT',' NULL')],
            # limite = registrosModificados,
            orden = ("debitos.id", "DESC")
        )

        ws['A1'] = 'Afiliado'
        ws['B1'] = 'Importe'
        ws['C1'] = 'Cuota'
        ws['D1'] = 'Cantidad de cuotas'
        ws['E1'] = 'Rechazado'
        ws['F1'] = 'Empresa'

        print("DBG - COSAS Proc: ", self.__debitosProcesados)
        print("DBG - COSAS Rech: ", self.__debitosRechazados)
        print("DBG - registros procesados: ", registrosProcesados)

        db_id_temporal, db_legajo = 6, 7
        fp_id_temporal, fp_legajo = 7, 2
        fr_id_temporal, fr_legajo = 1, 2

        index = 2
        for possMatch in self.__debitosProcesados:
            for debito in debitosXls:

                print(possMatch[fp_legajo], debito[db_legajo])
                print(possMatch[fp_id_temporal], debito[db_id_temporal])
                if possMatch[fp_legajo] == debito[db_legajo] and int(possMatch[fp_id_temporal]) == debito[db_id_temporal]:

                    a = 'A{}'.format(index)
                    b = 'B{}'.format(index)
                    c = 'C{}'.format(index)
                    d = 'D{}'.format(index)
                    e = 'E{}'.format(index)
                    f = 'F{}'.format(index)
                    ws[a] = debito[0]
                    ws[b] = debito[1]
                    ws[c] = debito[2]
                    ws[d] = debito[3]
                    ws[f] = debito[5]

                    print(debito, index)
                    index += 1

        index = 3
        for possMatch in self.__debitosRechazados:
            for debito in debitosXls:
                if int(possMatch[fr_id_temporal]) == debito[db_id_temporal] and possMatch[fr_legajo] == debito[db_legajo]:
                    a = 'A{}'.format(index + registrosProcesados)
                    b = 'B{}'.format(index + registrosProcesados)
                    c = 'C{}'.format(index + registrosProcesados)
                    d = 'D{}'.format(index + registrosProcesados)
                    e = 'E{}'.format(index + registrosProcesados)
                    f = 'F{}'.format(index + registrosProcesados)
                    ws[a] = debito[0]
                    ws[b] = debito[1]
                    ws[c] = debito[2]
                    ws[d] = debito[3]
                    ws[e] = self.__codigosDeRechazo[possMatch[13]]
                    ws[f] = debito[5]

                    print(debito, index)
                    index += 1

		# ABRIR UN CUADRO DE DIALOGO INDICANDO DONDE GUARDAR
        self.handleSave(wb)
        wb.close()

    def handleSave(self, workbook):
        path = QFileDialog.getSaveFileName(
        	None, 'Save File', '/comercios/', 'Excel(*.xlsx)')
        if not path[0]: return
        workbook.save(path[0])

    def limpiarTabla(self):
        self.__debitosBet = []
        self.layoutChanged.emit()

    def __setTotales(self, indexImporte):
        self.total_debitos = len(self.__debitosBet)
        self.importe_total = 0
        if self.total_debitos > 0:
            for debito in self.__debitosBet:
                self.importe_total += debito[indexImporte]

    def __toString(self, index):
        for debito in self.__debitosBet:
            debito[index] = str(debito[index])

    def validarPropiedades(self, propiedades):
# ahora mi función se asegura que las propieades existan en la lista, debería encontrar si hay alguna forma mas elegante de hacer esto
        if propiedades:
            prop = []
            for propiedad in propiedades:
                if propiedad in self.__propiedades:
                    prop.append(propiedad)
                else:
                    print("Propiedad '{}' es inválida, no se agregará".format(propiedad))
            return prop

# Estas son las funciones específicas de Qt para las tablas
    def rowCount(self, parent):
        return len(self.__debitosBet)

    def columnCount(self, parent):
        if self.__debitosBet:
            return len(self.__debitosBet[0])
        else:
            return 0

    def flags(self, index):
        return QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled

    def data(self, index, role):
# Acá es donde definí de dónde (De qué lista) voy a levantar los datos
        if role == QtCore.Qt.DisplayRole:
            row = index.row()
            column = index.column()
            value = self.__debitosBet[row][column] # value contiene la lista de listas que contiene los afiliados

            return value # el valor que retorno es el que aparecería en la tabla

    def setData(self, index, value, role = QtCore.Qt.EditRole):
# Esta función no la estoy usando
        if role == QtCore.Qt.EditRole:
            row = index.row()
            column = index.column()

            value = self.articulos[row][column]

            return value

    def headerData(self, section, orientation, role):
        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
# Los objetos tipo diccionario no ordenan sus elementos, por eso usar dict.keys() me tira los nombres
# de las columnas en cualquier orden. Acá debería usar la lista propiedades.
# además de salir ordenado, se ajusta la cantidad de columnas correspondiente
                keys = list(self.__propiedades)
                return keys[section]
